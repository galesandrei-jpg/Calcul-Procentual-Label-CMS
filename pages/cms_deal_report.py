import datetime as dt
import io
from typing import Dict, List

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from src.youtube import (
    YoutubeConfig,
    build_yta_service,
    list_group_items,
    get_channel_titles,
    query_channel_revenue_for_month,
)

st.set_page_config(page_title="CMS Deal Channel Report", layout="wide")

# Path to the template file (adjust if needed)
TEMPLATE_PATH = "template_raport_cms_deal.xlsx"


st.title("📋 CMS Deal – Channel Revenue Report")
st.caption(
    "Generates per-channel revenue report for a YouTube CMS group. "
    "Fetches channel IDs from the group, retrieves revenue metrics, "
    "and outputs an XLSX file with formulas (columns I–M) from the template."
)

# ----- Configuration -----
with st.expander("1) Configuration", expanded=True):
    default_owner = st.secrets.get("youtube", {}).get("content_owner", "")
    default_on_behalf = st.secrets.get("youtube", {}).get("on_behalf_of_content_owner", "")

    col1, col2 = st.columns(2)
    with col1:
        content_owner = st.text_input(
            "YouTube CMS Content Owner ID",
            value=default_owner,
            key="cms_deal_owner",
        )
    with col2:
        on_behalf = st.text_input(
            "onBehalfOfContentOwner",
            value=default_on_behalf,
            key="cms_deal_on_behalf",
        )

# ----- Group ID + Month -----
with st.expander("2) Group & Month Selection", expanded=True):
    group_id = st.text_input(
        "Channel Group ID",
        value="",
        help="Paste the group ID from studio.youtube.com/group/<GROUP_ID>/analytics",
        key="cms_deal_group_id",
    )

    now = dt.date.today()
    col1, col2 = st.columns(2)
    with col1:
        report_year = st.selectbox(
            "Year",
            list(range(2015, now.year + 2)),
            index=list(range(2015, now.year + 2)).index(now.year),
            key="cms_deal_year",
        )
    with col2:
        report_month = st.selectbox(
            "Month",
            list(range(1, 13)),
            index=now.month - 1,
            key="cms_deal_month",
        )

# ----- Run -----
with st.expander("3) Generate Report", expanded=True):
    run = st.button("🚀 Generate Channel Revenue Report", type="primary", key="cms_deal_run")

    if run:
        # Validate inputs
        if not content_owner.strip():
            st.error("Please provide the Content Owner ID.")
            st.stop()
        if not group_id.strip():
            st.error("Please provide a Channel Group ID.")
            st.stop()

        try:
            ycfg = YoutubeConfig(
                content_owner=content_owner.strip(),
                on_behalf_of_content_owner=on_behalf.strip() or None,
            )
            yta = build_yta_service(ycfg)

            status = st.empty()
            progress = st.progress(0.0)

            # Step 1: Get channel IDs from the group
            status.info("Fetching channels from group…")
            group_items = list_group_items(yta, ycfg, group_id.strip())
            channel_ids = [item["channelId"] for item in group_items]

            if not channel_ids:
                st.error("No channels found in this group.")
                st.stop()

            st.success(f"Found {len(channel_ids)} channels in group.")
            progress.progress(0.15)

            # Step 2: Get channel titles
            status.info("Fetching channel titles…")
            titles_map = get_channel_titles(channel_ids)
            progress.progress(0.30)

            # Step 3: Query total revenue (all countries)
            status.info("Querying total revenue (all countries)…")
            total_revenue = query_channel_revenue_for_month(
                yta, ycfg,
                year=report_year,
                month=report_month,
                group_id=group_id.strip(),
                country=None,
            )
            progress.progress(0.60)

            # Step 4: Query US-only revenue
            status.info("Querying US-only revenue…")
            us_revenue = query_channel_revenue_for_month(
                yta, ycfg,
                year=report_year,
                month=report_month,
                group_id=group_id.strip(),
                country="US",
            )
            progress.progress(0.85)

            # Step 5: Build the data rows
            status.info("Building report…")
            rows_data = []
            for cid in channel_ids:
                title = titles_map.get(cid, cid)
                t = total_revenue.get(cid, {})
                u = us_revenue.get(cid, {})

                rows_data.append({
                    "Channel": cid,
                    "Channel title": title,
                    "Estimated partner revenue (USD)": t.get("estimatedPartnerRevenue", 0.0),
                    "Estimated partner ad revenue (USD)": t.get("estimatedPartnerAdRevenue", 0.0),
                    "YouTube Premium partner revenue (USD)": t.get("estimatedPartnerPremiumRevenue", 0.0),
                    "Total US": u.get("estimatedPartnerRevenue", 0.0),
                    "US Revenue Ad": u.get("estimatedPartnerAdRevenue", 0.0),
                    "US Revenue Premium": u.get("estimatedPartnerPremiumRevenue", 0.0),
                })

            df = pd.DataFrame(rows_data)

            # Sort by total revenue descending
            df = df.sort_values(
                "Estimated partner revenue (USD)", ascending=False
            ).reset_index(drop=True)

            progress.progress(0.90)

            # Step 6: Write into XLSX template
            status.info("Writing XLSX with formulas…")

            wb = load_workbook(TEMPLATE_PATH)
            ws = wb.active

            # Row 2 is the SUM row (formulas already in template)
            # Data starts at row 3
            DATA_START_ROW = 3
            num_channels = len(df)
            data_end_row = DATA_START_ROW + num_channels - 1

            # Write data columns A–H (columns 1–8)
            col_mapping = {
                1: "Channel",
                2: "Channel title",
                3: "Estimated partner revenue (USD)",
                4: "Estimated partner ad revenue (USD)",
                5: "YouTube Premium partner revenue (USD)",
                6: "Total US",
                7: "US Revenue Ad",
                8: "US Revenue Premium",
            }

            for i, (_, row_data) in enumerate(df.iterrows()):
                excel_row = DATA_START_ROW + i
                for col_num, col_name in col_mapping.items():
                    value = row_data[col_name]
                    ws.cell(row=excel_row, column=col_num, value=value)

                # Write formulas for columns I–M
                r = excel_row
                ws.cell(row=r, column=9, value=f"=F{r}*0.1")        # I: US Tax
                ws.cell(row=r, column=10, value=f"=C{r}-I{r}")       # J: Gross Partner Revenue
                ws.cell(row=r, column=11, value=f"=J{r}*0.85")       # K: NET Partner Revenue EUR
                ws.cell(row=r, column=12, value=f"=K{r}*13%")        # L: Gross HaHaHa Partner Revenue
                ws.cell(row=r, column=13, value=f"=K{r}-L{r}")       # M: Gross Partner Revenue

            # Update row 2 SUM formulas to cover actual data range
            ws.cell(row=2, column=9, value=f"=SUM(I{DATA_START_ROW}:I{data_end_row})")
            ws.cell(row=2, column=10, value=f"=SUM(J{DATA_START_ROW}:J{data_end_row})")
            ws.cell(row=2, column=11, value=f"=SUM(K{DATA_START_ROW}:K{data_end_row})")
            ws.cell(row=2, column=12, value=f"=K2*13%")
            ws.cell(row=2, column=13, value=f"=K2-L2")

            # Also add SUM formulas for columns A–H numeric columns (C–H) in row 2
            for col_num in range(3, 9):
                ws.cell(
                    row=2, column=col_num,
                    value=f"=SUM({get_column_letter(col_num)}{DATA_START_ROW}:{get_column_letter(col_num)}{data_end_row})"
                )

            # Save to BytesIO
            output_buf = io.BytesIO()
            wb.save(output_buf)
            output_buf.seek(0)
            xlsx_bytes = output_buf.getvalue()

            progress.progress(1.0)
            status.success(f"Done ✅ Report generated with {num_channels} channels.")

            # Display preview
            st.subheader("Preview")
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Channels", num_channels)
            with col2:
                total_rev = df["Estimated partner revenue (USD)"].sum()
                st.metric("Total Revenue (USD)", f"${total_rev:,.2f}")
            with col3:
                total_us = df["Total US"].sum()
                st.metric("Total US Revenue (USD)", f"${total_us:,.2f}")

            st.dataframe(df, use_container_width=True)

            # Download
            month_str = f"{report_year:04d}-{report_month:02d}"
            filename = f"cms_deal_report_{month_str}.xlsx"
            st.download_button(
                label="📥 Download XLSX Report",
                data=xlsx_bytes,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.spreadsheet",
            )

        except Exception as e:
            st.exception(e)
